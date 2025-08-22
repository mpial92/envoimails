import openpyxl
import smtplib
import os
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Charger le fichier Excel
chemin_fichier_excel = 'liste.xlsx'
wb = openpyxl.load_workbook(chemin_fichier_excel)
sheet = wb.active

# Paramètres du compte et du serveur SMTP depuis les variables d'environnement
adresse_email = os.getenv("EMAIL_USERNAME", "contact@lecercledesseniorsneuflizeobc.fr")
mot_de_passe = os.getenv("EMAIL_PASSWORD")
serveur_smtp = os.getenv("SMTP_SERVER", "ssl0.ovh.net")
port_smtp = int(os.getenv("SMTP_PORT", "587"))

# Vérification des variables d'environnement critiques
if not mot_de_passe:
    print("ERREUR: Le mot de passe email n'est pas configuré dans les variables d'environnement")
    exit(1)

print(f"Configuration SMTP: {serveur_smtp}:{port_smtp}")
print(f"Adresse email: {adresse_email}")

compteur_emails = 0

# Fonction pour se connecter au serveur SMTP
def connecter_au_serveur_smtp():
    try:
        server = smtplib.SMTP(serveur_smtp, port_smtp)
        server.starttls()
        server.login(adresse_email, mot_de_passe)
        print("Connexion SMTP réussie")
        return server
    except Exception as e:
        print(f"Erreur de connexion SMTP: {e}")
        return None

# Connexion initiale au serveur SMTP
server = connecter_au_serveur_smtp()
if not server:
    print("Impossible de se connecter au serveur SMTP")
    exit(1)

# Parcourir les lignes du fichier Excel
print(f"Nombre total de lignes à traiter: {sheet.max_row - 1}")

for row in range(2, sheet.max_row + 1):
    destinataire = sheet.cell(row=row, column=12).value
    nom = sheet.cell(row=row, column=2).value
    
    # Vérifier que l'adresse email existe
    if not destinataire:
        print(f"Ligne {row}: Pas d'adresse email, passage à la suivante")
        continue

    # Créer le message
    msg = MIMEMultipart()
    msg['From'] = adresse_email
    msg['To'] = destinataire
    msg['Subject'] = "Nouvelles"

    # Corps du message
    body = """
<html>
  <body>
    <p>Bonjour à toutes et à tous,</p>

  </body>
</html>
"""
 
    msg.attach(MIMEText(body, 'html'))

    # Envoyer les emails 
    try:
        # ATTENTION: Ligne suivante commentée pour les tests - décommente pour l'envoi réel
        #server.sendmail(adresse_email, destinataire, msg.as_string())
        print(f"E-mail envoyé à {destinataire} (nom: {nom})")
        compteur_emails += 1 

        # Pause toutes les 10 e-mails
        if compteur_emails % 10 == 0:
            print("Pause de 1 minute pour éviter les filtres anti-spam...")
            time.sleep(60)  # Pause de 1 minute
            server.quit()  # Fermer la connexion actuelle
            server = connecter_au_serveur_smtp()  # Se reconnecter au serveur SMTP
            if not server:
                print("Erreur de reconnexion SMTP")
                break
    
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'e-mail à {destinataire} : {e}")

# Déconnexion du serveur SMTP
if server:
    server.quit()
    print(f"Workflow terminé. Total d'emails traités: {compteur_emails}")
else:
    print("Aucune connexion SMTP à fermer")